"""
Catering Costing Dashboard — ETL Script
Transforms "Copy of Costing Schedule.xlsx" into a clean relational model
with two sheets: Dim_Menu and Fact_Ingredients, ready for Power BI import.

Requirements:  pip install pandas openpyxl
Usage:         python catering_etl.py
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Paths ─────────────────────────────────────────────────────────────────────
INPUT_FILE  = "Copy of Costing Schedule.xlsx"   # adjust if needed
OUTPUT_FILE = "Catering_Model.xlsx"

# ══════════════════════════════════════════════════════════════════════════════
# PART 1A  —  Dim_Menu  (from 'Menu Cost Summary')
# ══════════════════════════════════════════════════════════════════════════════
def build_dim_menu(xl: pd.ExcelFile) -> pd.DataFrame:
    """
    Load Menu Cost Summary and return a clean Dim_Menu table.
    Columns: [Serial No, Menu Name, Raw Cost, Overhead Cost, Total Cost, Final Value]
    """
    raw = pd.read_excel(xl, sheet_name="Menu Cost Summary", header=None)

    # Header is always at row-index 4; data starts at row-index 5
    df = raw.iloc[5:, [0, 1, 2, 4, 5, 7]].copy()
    df.columns = ["Serial No", "Menu Name", "Raw Cost",
                  "Overhead Cost", "Total Cost", "Final Value"]

    # Keep only true data rows (Serial No must be numeric)
    df = df[pd.to_numeric(df["Serial No"], errors="coerce").notna()].copy()
    df = df.dropna(subset=["Menu Name"]).reset_index(drop=True)

    for col in ["Serial No", "Raw Cost", "Overhead Cost", "Total Cost", "Final Value"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    return df


# ══════════════════════════════════════════════════════════════════════════════
# PART 1B  —  Fact_Ingredients  (from 'Menu Cost Details')
# ══════════════════════════════════════════════════════════════════════════════
def build_fact_ingredients(xl: pd.ExcelFile) -> pd.DataFrame:
    """
    Parse the raw recipe sheet with nested recipe blocks.
    Each block has:
      • A header row: Serial No | Menu Name | first ingredient | …
      • A batch-size row: NaN | <number> | next ingredient | …
      • A 'Pax' label row: NaN | 'Pax' | next ingredient | …
      • Plain ingredient rows: NaN | NaN | ingredient | …
      • Subtotal / separator rows (no Ingredients value) — skipped

    Returns Fact_Ingredients with columns:
      [Menu Name, Ingredients, Unit, Qty Per Pax, Price]
    """
    raw = pd.read_excel(xl, sheet_name="Menu Cost Details", header=None)
    data = raw.iloc[5:].copy()
    data.columns = ["Serial_No", "Col1", "Ingredients", "Unit", "Qty", "Price",
                    "Total_Value", "_X", "_Y"]
    data = data.reset_index(drop=True)

    records   = []
    menu_name = None
    batch_pax = None

    for _, row in data.iterrows():
        sn   = row["Serial_No"]
        c1   = row["Col1"]
        ingr = row["Ingredients"]
        unit = row["Unit"]
        qty  = row["Qty"]
        price = row["Price"]

        c1_str = str(c1).strip() if pd.notna(c1) else ""

        # ── New recipe block: Serial_No is a valid integer ────────────────
        if pd.notna(sn):
            try:
                int(float(str(sn).strip()))
                menu_name = c1_str if c1_str not in ("", "nan") else menu_name
                batch_pax = None          # reset; will be detected in next rows
            except ValueError:
                pass

        # ── Detect batch-pax: Col1 is numeric and Serial_No is blank ──────
        if pd.isna(sn) and c1_str not in ("", "nan", "Pax"):
            try:
                bp = float(c1_str)
                if bp > 0 and batch_pax is None:
                    batch_pax = bp
            except ValueError:
                pass

        # ── Collect ingredient row ─────────────────────────────────────────
        if pd.notna(ingr) and str(ingr).strip() not in ("", "nan"):
            qty_val   = pd.to_numeric(qty,   errors="coerce")
            price_val = pd.to_numeric(price, errors="coerce")
            records.append({
                "Menu Name":  menu_name,
                "Ingredients": str(ingr).strip(),
                "Unit":        str(unit).strip() if pd.notna(unit) else "",
                "Qty":         qty_val,
                "Batch Pax":   batch_pax,
                "Price":       price_val,
            })

    fact = pd.DataFrame(records)

    # ── Filter out rows with missing critical values ───────────────────────
    fact = fact.dropna(subset=["Menu Name", "Qty", "Batch Pax"])
    fact = fact[fact["Qty"] > 0].copy()

    # ── STEP 6: Normalise Qty to per-person basis ──────────────────────────
    fact["Qty Per Pax"] = (fact["Qty"] / fact["Batch Pax"]).round(6)

    return fact[["Menu Name", "Ingredients", "Unit", "Qty Per Pax", "Price"]].reset_index(drop=True)


# ══════════════════════════════════════════════════════════════════════════════
# WRITE TO EXCEL  —  professional formatting with openpyxl
# ══════════════════════════════════════════════════════════════════════════════
HEADER_COLOR = "2F5496"    # dark blue
ALT_ROW_COLOR = "EEF2FF"  # light lavender


def _apply_header(ws, headers: list[str]) -> None:
    fill   = PatternFill("solid", start_color=HEADER_COLOR, end_color=HEADER_COLOR)
    font   = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(1, col, h)
        cell.fill  = fill
        cell.font  = font
        cell.alignment = align
    ws.row_dimensions[1].height = 28


def _apply_data_style(ws, n_rows: int, n_cols: int) -> None:
    alt    = PatternFill("solid", start_color=ALT_ROW_COLOR, end_color=ALT_ROW_COLOR)
    side   = Side(style="thin", color="D0D0D0")
    border = Border(left=side, right=side, top=side, bottom=side)
    for r in range(2, n_rows + 2):
        for c in range(1, n_cols + 1):
            cell = ws.cell(r, c)
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center")
            cell.border    = border
            if r % 2 == 0:
                cell.fill = alt


def write_excel(dim_menu: pd.DataFrame, fact_ingredients: pd.DataFrame,
                output_path: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Dim_Menu ─────────────────────────────────────────────────
    ws1 = wb.create_sheet("Dim_Menu")
    headers1 = ["Serial No", "Menu Name", "Raw Cost",
                "Overhead Cost", "Total Cost", "Final Value"]
    _apply_header(ws1, headers1)

    for r_idx, row in dim_menu.iterrows():
        for c_idx, h in enumerate(headers1, 1):
            val = row[h]
            ws1.cell(r_idx + 2, c_idx, None if pd.isna(val) else val)

    _apply_data_style(ws1, len(dim_menu), len(headers1))
    ws1.column_dimensions["A"].width = 12
    ws1.column_dimensions["B"].width = 34
    for col in ["C", "D", "E", "F"]:
        ws1.column_dimensions[col].width = 16

    # ── Sheet 2: Fact_Ingredients ─────────────────────────────────────────
    ws2 = wb.create_sheet("Fact_Ingredients")
    headers2 = ["Menu Name", "Ingredients", "Unit", "Qty Per Pax", "Price"]
    _apply_header(ws2, headers2)

    for r_idx, row in fact_ingredients.reset_index(drop=True).iterrows():
        for c_idx, h in enumerate(headers2, 1):
            val = row[h]
            if isinstance(val, float):
                val = round(val, 6)
            ws2.cell(r_idx + 2, c_idx, None if (isinstance(val, float) and np.isnan(val)) else val)

    _apply_data_style(ws2, len(fact_ingredients), len(headers2))
    ws2.column_dimensions["A"].width = 32
    ws2.column_dimensions["B"].width = 30
    ws2.column_dimensions["C"].width = 10
    ws2.column_dimensions["D"].width = 14
    ws2.column_dimensions["E"].width = 12

    wb.save(output_path)
    print(f"✅  Saved: {output_path}")
    print(f"   Dim_Menu         : {len(dim_menu):>5} rows")
    print(f"   Fact_Ingredients : {len(fact_ingredients):>5} rows")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    xl = pd.ExcelFile(INPUT_FILE)

    dim_menu         = build_dim_menu(xl)
    fact_ingredients = build_fact_ingredients(xl)

    write_excel(dim_menu, fact_ingredients, OUTPUT_FILE)
